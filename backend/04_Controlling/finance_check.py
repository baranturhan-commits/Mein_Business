import os
import sys
import json
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# Force UTF-8 for Console
sys.stdout.reconfigure(encoding='utf-8')

# --- CONFIG & PATHS ---
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")
AUSGABEN_BASE_DIR = os.path.join(BASE_DIR, "02_Buchhaltung", "Ausgaben")
EINNAHMEN_FILE = os.path.join(BASE_DIR, "03_Rechnungen", "Einnahmen.csv")

def clear_screen():
    os.system('cls' if os.name == 'nt' else 'clear')

# --- HELPERS ---
def load_profiles():
    if not os.path.exists(CONFIG_FILE):
        return []
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
            return data.get('profiles', [])
    except:
        return []

def clean_currency(val):
    if pd.isna(val) or val == '': return 0.0
    s = str(val).replace('€', '').replace('EUR', '').strip()
    s = s.replace('.', '') 
    s = s.replace(',', '.') 
    try:
        return float(s)
    except:
        return 0.0

def parse_german_date(d_str):
    try:
        return pd.to_datetime(d_str, format='%d.%m.%Y', dayfirst=True)
    except:
        return pd.NaT

# --- EXCEL STYLING HELPERS ---
def auto_size_columns(ws):
    """Passt Spaltenbreite automatisch an (Robust gegen MergedCells)."""
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        # Iteriere über alle Zellen in dieser Spalte
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except:
                pass
                
        adjusted_width = (max_length + 2)
        if adjusted_width < 25:
            adjusted_width = 25
            
        try:
            ws.column_dimensions[column_letter].width = adjusted_width
        except:
            pass

def style_standard_cell(cell, bold=False, align_right=False, bg_color=None):
    cell.font = Font(name='Calibri', size=12, bold=bold)
    if align_right:
        cell.alignment = Alignment(horizontal='right')
    else:
        cell.alignment = Alignment(horizontal='left')
    
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
        
    cell.border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

# --- MAIN ---
def main():
    clear_screen()
    print("📊 FINANCE CHECK & DASHBOARD")
    print("="*60)

    # 1. SETUP & SELECTION
    profiles = load_profiles()
    if not profiles:
        print("❌ Keine Firmenprofile in config.json gefunden.")
        return

    print("Für welche Firma wollen wir die Finanzen prüfen?")
    for idx, p in enumerate(profiles, 1):
        print(f"[{idx}] {p.get('firma', 'Unbekannt')}")
    
    try:
        c_idx = int(input("\nAuswahl: ")) - 1
        if c_idx < 0 or c_idx >= len(profiles): raise ValueError
        current_profile = profiles[c_idx]
    except:
        print("❌ Ungültige Auswahl.")
        return

    firma_name = current_profile['firma']
    firma_safe = "".join(c for c in firma_name if c.isalnum() or c in ('_', '-')).strip()
    print(f"\n✅ Gewählt: {firma_name}")

    print("\nWelcher Zeitraum?")
    print("[1] Aktueller Monat")
    print("[2] Letzter Monat")
    print("[3] Ganzes Jahr (2025)")
    print("[4] Alles (Gesamt)")
    
    time_choice = input("\nAuswahl: ").strip()
    
    # 2. LOAD DATA
    print("\n🔄 Lade Daten...")
    
    # A) Ausgaben
    ausgaben_path = os.path.join(AUSGABEN_BASE_DIR, firma_safe, "Ausgaben.csv")
    df_aus = pd.DataFrame()
    if os.path.exists(ausgaben_path):
        try:
            df_aus = pd.read_csv(ausgaben_path, sep=';', encoding='utf-8-sig')
            for c in ['Netto', 'MwSt', 'Brutto']:
                if c in df_aus.columns: df_aus[c] = df_aus[c].apply(clean_currency)
            df_aus['Datum_DT'] = df_aus['Datum'].apply(parse_german_date)
        except Exception as e:
            print(f"⚠️  Fehler bei Ausgaben lesen: {e}")

    # B) Einnahmen
    df_ein = pd.DataFrame()
    if os.path.exists(EINNAHMEN_FILE):
        try:
            df_ein = pd.read_csv(EINNAHMEN_FILE, sep=';', encoding='utf-8-sig', on_bad_lines='skip')
            for c in ['Netto', 'MwSt', 'Brutto']:
                if c in df_ein.columns: df_ein[c] = df_ein[c].apply(clean_currency)
            df_ein['Datum_DT'] = df_ein['Datum'].apply(parse_german_date)
            # Filter Company
            for col in ['Firma', 'Absender', 'Company']:
                if col in df_ein.columns:
                    df_ein = df_ein[df_ein[col] == firma_name]
                    break
        except Exception as e:
            print(f"⚠️  Fehler bei Einnahmen lesen: {e}")

    # 3. FILTER TIMEFRAME
    now = datetime.date.today()
    start_date = None
    end_date = None
    
    if time_choice == '1': # This Month
        start_date = now.replace(day=1)
    elif time_choice == '2': # Last Month
        last_month_end = now.replace(day=1) - datetime.timedelta(days=1)
        start_date = last_month_end.replace(day=1)
        end_date = last_month_end
    elif time_choice == '3': # 2025
        start_date = datetime.date(2025, 1, 1)
        end_date = datetime.date(2025, 12, 31)
    
    def filter_df(df):
        if df.empty or 'Datum_DT' not in df.columns: return df
        mask = pd.Series([True]*len(df))
        if start_date:
            mask = mask & (df['Datum_DT'].dt.date >= start_date)
        if end_date:
            mask = mask & (df['Datum_DT'].dt.date <= end_date)
        if time_choice == '1' and not end_date:
             mask = mask & (df['Datum_DT'].dt.month == start_date.month) & (df['Datum_DT'].dt.year == start_date.year)
        return df[mask]

    df_aus = filter_df(df_aus)
    df_ein = filter_df(df_ein)

    # 4. CALCULATIONS
    ein_netto = df_ein['Netto'].sum() if not df_ein.empty and 'Netto' in df_ein.columns else 0.0
    ein_mwst  = df_ein['MwSt'].sum() if not df_ein.empty and 'MwSt' in df_ein.columns else 0.0
    
    aus_netto = df_aus['Netto'].sum() if not df_aus.empty and 'Netto' in df_aus.columns else 0.0
    aus_mwst  = df_aus['MwSt'].sum() if not df_aus.empty and 'MwSt' in df_aus.columns else 0.0
    
    gewinn = ein_netto - aus_netto
    zahllast = ein_mwst - aus_mwst
    
    # --- DEUTSCHE STEUER-LOGIK ---
    
    # A) Gewerbesteuer
    # Freibetrag 24.500 EUR für Einzelunternehmen / Personengesellschaften.
    # Ansonsten grob 3,5% Steuermesszahl * Hebesatz (oft 300-400%). ~14-15%.
    gwst = 0.0
    if gewinn > 24500:
        gwst = (gewinn - 24500) * 0.15 
    
    # B) Einkommensteuer (Dynamisch mit Grundfreibetrag)
    # 2024/2025 Grundfreibetrag ca 11.604 Euro.
    est_satz = 0.0
    if gewinn < 11604:
        est_satz = 0.0 # Steuerfrei
    elif gewinn < 60000:
        est_satz = 0.30 # Progressiver Mittelwert Puffer
    else:
        est_satz = 0.42 # Spitzensteuersatz
        
    est_ruecklage = gewinn * est_satz
    
    # C) Verfügbar
    reingewinn = gewinn - gwst - est_ruecklage

    # 5. TERMINAL OUTPUT
    print("\n" + "="*40)
    print(f"📈 FINANZ-ÜBERSICHT: {firma_name}")
    print("="*40)
    print(f"➕ Einnahmen (Netto): {ein_netto:10.2f} €")
    print(f"➖ Ausgaben (Netto) : {aus_netto:10.2f} €")
    print("-" * 40)
    print(f"💰 GEWINN (v.St.)   : {gewinn:10.2f} €")
    print("-" * 40)
    icon_tax = "🟢" if zahllast < 0 else "🔴"
    print(f"{icon_tax} Umsatzsteuer     : {zahllast:10.2f} €")
    
    if gwst > 0:
        print(f"� Gewerbesteuer    : {gwst:10.2f} € (Freigrenze überschritten)")
    else:
        print(f"🏭 Gewerbesteuer    : {gwst:10.2f} € (unter 24.500€)")
        
    print(f"🏛️ Einkommensteuer  : {est_ruecklage:10.2f} € (Satz ca. {int(est_satz*100)}%)")
    print("-" * 40)
    print(f"💵 VERFÜGBAR        : {reingewinn:10.2f} € (Privat)")

    # 6. EXCEL EXPORT (STYLED)
    report_folder = os.path.join(SCRIPT_DIR, "Reports")
    if not os.path.exists(report_folder): os.makedirs(report_folder)
    
    t_label = "Gesamt"
    if time_choice == '1': t_label = "AktuellerMonat"
    elif time_choice == '2': t_label = "LetzterMonat"
    elif time_choice == '3': t_label = "2025"
    
    filename = f"{firma_safe}_{t_label}_Report.xlsx"
    filepath = os.path.join(report_folder, filename)
    
    wb = Workbook()
    
    # --- SHEET 1: DASHBOARD ---
    ws1 = wb.active
    ws1.title = "Dashboard"
    
    # Title
    ws1.merge_cells('A1:C1')
    title_cell = ws1['A1']
    title_cell.value = f"FINANZ-REPORT: {firma_name}"
    title_cell.font = Font(name='Calibri', size=24, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color="00008B", end_color="00008B", fill_type="solid") # DarkBlue
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    rows_data = [
        ("KENNZAHLEN", "", "CCCCCC", "header"),
        ("Einnahmen Netto", ein_netto, "90EE90", "income"), # LightGreen
        ("Ausgaben Netto", aus_netto, "FFC0CB", "expense"), # Pink/LightRed
        ("", "", "FFFFFF", "spacer"),
        ("GEWINN (Operativ)", gewinn, "FFD700", "profit"), # Gold
        ("", "", "FFFFFF", "spacer"),
        ("STEUERN & ABGABEN", "", "CCCCCC", "header"),
        ("USt Zahllast (Finanzamt)", zahllast, "D3D3D3", "tax"),
        ("Gewerbesteuer (ab 24.5k)", gwst, "D3D3D3", "tax"),
        (f"Einkommensteuer ({int(est_satz*100)}%)", est_ruecklage, "D3D3D3", "tax"),
        ("", "", "FFFFFF", "spacer"),
        ("VERFÜGBARER GEWINN (Privat)", reingewinn, "00B050", "final") # Bold Green
    ]
    
    start_row = 3
    for idx, (label, val, bg, style) in enumerate(rows_data):
        r = start_row + idx
        c1 = ws1.cell(row=r, column=1, value=label)
        c2 = ws1.cell(row=r, column=2, value=val)
        
        # Base Style
        style_standard_cell(c1, bg_color=bg)
        style_standard_cell(c2, bg_color=bg, align_right=True)
        
        # Specifics
        if style == 'header':
            c1.font = Font(name='Calibri', size=14, bold=True)
            c2.value = ""
        elif style == 'profit':
            profit_font = Font(name='Calibri', size=16, bold=True)
            c1.font = profit_font
            c2.font = profit_font
            double_border = Side(style='double')
            border = Border(left=double_border, right=double_border, top=double_border, bottom=double_border)
            c1.border = border
            c2.border = border
        elif style == 'final':
            final_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF') # White Text
            c1.font = final_font
            c2.font = final_font
            thick_border_side = Side(style='thick') # Thick Border
            thick_border = Border(left=thick_border_side, right=thick_border_side, top=thick_border_side, bottom=thick_border_side)
            c1.border = thick_border
            c2.border = thick_border
            
        if isinstance(val, (int, float)):
            c2.number_format = '#,##0.00 €'

    auto_size_columns(ws1)

    # --- SHEET 2: EINNAHMEN ---
    ws2 = wb.create_sheet("Einnahmen")
    if not df_ein.empty:
        if 'Datum_DT' in df_ein.columns: df_ein = df_ein.drop(columns=['Datum_DT'])
        headers = list(df_ein.columns)
        ws2.append(headers)
        for cell in ws2[1]:
            style_standard_cell(cell, bold=True, bg_color="90EE90")
        for r_idx, row in enumerate(dataframe_to_rows(df_ein, index=False, header=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws2.cell(row=r_idx, column=c_idx, value=val)
                is_num = isinstance(val, (int, float))
                style_standard_cell(cell, align_right=is_num)
                if is_num and c_idx > 1:
                     cell.number_format = '#,##0.00 €'
    auto_size_columns(ws2)

    # --- SHEET 3: AUSGABEN ---
    ws3 = wb.create_sheet("Ausgaben")
    if not df_aus.empty:
        if 'Datum_DT' in df_aus.columns: df_aus = df_aus.drop(columns=['Datum_DT'])
        headers = list(df_aus.columns)
        ws3.append(headers)
        for cell in ws3[1]:
            style_standard_cell(cell, bold=True, bg_color="FFC0CB")
        for r_idx, row in enumerate(dataframe_to_rows(df_aus, index=False, header=False), 2):
            for c_idx, val in enumerate(row, 1):
                cell = ws3.cell(row=r_idx, column=c_idx, value=val)
                is_num = isinstance(val, (int, float))
                style_standard_cell(cell, align_right=is_num)
                if is_num and c_idx > 1:
                     cell.number_format = '#,##0.00 €'
    auto_size_columns(ws3)

    wb.save(filepath)
    print(f"\n✅ Excel-Report erstellt: {filepath}")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n👋 Abbruch.")
    except Exception as e:
        print(f"\n❌ Fehler: {e}")

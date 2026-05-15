import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --- STYLING CONSTANTS ---
HEADER_FONT = Font(name='Calibri', size=11, bold=True)
BODY_FONT = Font(name='Calibri', size=11)
BORDER_THIN = Side(style='thin')
BORDER_SET = Border(left=BORDER_THIN, right=BORDER_THIN, top=BORDER_THIN, bottom=BORDER_THIN)

THEME_COLORS = {
    'Einnahmen': '90EE90', # LightGreen
    'Ausgaben': 'FFC0CB',  # Pink
    'Default': 'D3D3D3'    # LightGray
}

def auto_size_columns(ws):
    """Passt Spaltenbreite automatisch an."""
    for col_idx in range(1, ws.max_column + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    val_len = len(str(cell.value))
                    if val_len > max_length:
                        max_length = val_len
            except: pass
        
        adjusted_width = (max_length + 2)
        # Limit width
        if adjusted_width > 50: adjusted_width = 50
        if adjusted_width < 10: adjusted_width = 10
            
        try:
            ws.column_dimensions[column_letter].width = adjusted_width
        except: pass

def style_cell(cell, is_header=False, bg_color=None, is_currency=False):
    cell.font = HEADER_FONT if is_header else BODY_FONT
    cell.border = BORDER_SET
    
    if is_header:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if bg_color:
             cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    else:
        # Body
        if is_currency:
             cell.alignment = Alignment(horizontal='right')
             cell.number_format = '#,##0.00 €'
        else:
             cell.alignment = Alignment(horizontal='left')

def init_file(filepath, headers, sheet_name="Sheet1"):
    """Erstellt eine neue leere Excel-Datei mit Headern."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    
    # Ermittle Farbe basierend auf Sheetname (grob)
    color = THEME_COLORS.get(sheet_name, THEME_COLORS['Default'])
    if "Ausgaben" in sheet_name: color = THEME_COLORS['Ausgaben']
    if "Einnahmen" in sheet_name: color = THEME_COLORS['Einnahmen']
    
    for cell in ws[1]:
        style_cell(cell, is_header=True, bg_color=color)
        
    wb.save(filepath)
    return wb

def append_data(filepath, data_dict, sheet_name="Data", headers=None):
    """
    Fügt eine Zeile (Dict) in die Excel-Datei ein.
    Erstellt Datei neu, falls sie nicht existiert.
    Liest existierende Header, um Reihenfolge zu wahren.
    """
    if headers is None:
        headers = list(data_dict.keys())

    # 1. Datei laden oder erstellen
    if not os.path.exists(filepath):
        init_file(filepath, headers, sheet_name)

    wb = load_workbook(filepath)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
        ws.append(headers)
        # Style Header New Sheet
        color = THEME_COLORS.get('Default')
        if "Ausgaben" in sheet_name: color = THEME_COLORS['Ausgaben']
        if "Einnahmen" in sheet_name: color = THEME_COLORS['Einnahmen']
        for cell in ws[1]:
             style_cell(cell, is_header=True, bg_color=color)

    # 2. Bestehende Header lesen
    existing_headers = [cell.value for cell in ws[1]]
    
    # Neue Header anfügen, falls im Parameter 'headers' welche fehlen
    if headers:
        for h in headers:
            if h not in existing_headers:
                ws.cell(row=1, column=len(existing_headers)+1, value=h)
                # Style New Header
                new_cell = ws.cell(row=1, column=len(existing_headers)+1)
                color = THEME_COLORS.get('Default')
                if "Ausgaben" in sheet_name: color = THEME_COLORS['Ausgaben']
                if "Einnahmen" in sheet_name: color = THEME_COLORS['Einnahmen']
                style_cell(new_cell, is_header=True, bg_color=color)
                
                existing_headers.append(h)
    # Mapping
    row_values = []
    for h in existing_headers:
        val = data_dict.get(h, "")
        # Versuch Zahlen zu konvertieren für Excel
        try:
            if isinstance(val, str) and ('€' in val or ',' in val):
                clean = val.replace('€','').replace('.','').replace(',','.').strip()
                val = float(clean)
        except: pass
        row_values.append(val)
        
    # 3. Append Row
    ws.append(row_values)
    
    # 4. Style New Row
    last_row_idx = ws.max_row
    for col_idx, cell in enumerate(ws[last_row_idx], 1):
        # Check Currency Context (Simple heuristic: Header contains 'Betrag', 'Netto', 'Brutto', 'MwSt')
        header_val = existing_headers[col_idx-1] if col_idx-1 < len(existing_headers) else ""
        is_cur = any(k in header_val for k in ['Betrag', 'Netto', 'Brutto', 'MwSt'])
        style_cell(cell, is_currency=is_cur)

    # 5. Fix Widths (Optional: kann langsam sein bei großen Files, hier ok)
    auto_size_columns(ws)
    
    wb.save(filepath)
    wb.close()
    return True

_CACHE = {}

def read_data(filepath, sheet_name=0):
    """Liest Daten via Pandas als Liste von Dicts (mit Cache für bessere Performance)."""
    if not os.path.exists(filepath):
        return []
        
    # Check cache based on file modification time
    try:
        mtime = os.path.getmtime(filepath)
        cache_key = f"{filepath}_{sheet_name}"
        if cache_key in _CACHE and _CACHE[cache_key]['mtime'] == mtime:
            return _CACHE[cache_key]['data']
    except Exception:
        pass
        
    try:
        df = pd.read_excel(filepath, sheet_name=sheet_name)
        # Robust cleanup: Cast to object to allow None everywhere
        df = df.astype(object)
        # Replace all pandas/numpy NaNs with None
        df = df.where(pd.notnull(df), None)
        data = df.to_dict('records')
        
        # Save to cache
        try:
            _CACHE[cache_key] = {'mtime': mtime, 'data': data}
        except Exception:
            pass
            
        return data
    except Exception as e:
        print(f"⚠️  Fehler beim Lesen von Excel {filepath}: {e}")
        return []

def update_status(filepath, id_col, id_val, status_col, new_status):
    """
    Aktualisiert einen Status-Wert in der Excel Datei unter Beibehaltung des Stylings.
    Nutzt openpyxl direkt.
    """
    if not os.path.exists(filepath): return False
    
    try:
        wb = load_workbook(filepath)
        ws = wb.active # Default sheet
        
        # Header Index finden
        header_row = [c.value for c in ws[1]]
        try:
            col_idx_id = header_row.index(id_col) + 1
            col_idx_stat = header_row.index(status_col) + 1
        except ValueError:
            print(f"❌ Spalten nicht gefunden: {id_col} oder {status_col}")
            return False
            
        found = False
        for row in ws.iter_rows(min_row=2):
            cell_id = row[col_idx_id-1]
            if str(cell_id.value) == str(id_val):
                cell_stat = row[col_idx_stat-1]
                cell_stat.value = new_status
                found = True
                # Style Update nicht nötig, behält altes, aber wir stellen sicher
                style_cell(cell_stat, is_header=False) 
                break
                
        if found:
            wb.save(filepath)
            wb.close()
            return True
        else:
            return False
            
    except Exception as e:
        print(f"❌ Excel Update Fehler: {e}")
        return False
